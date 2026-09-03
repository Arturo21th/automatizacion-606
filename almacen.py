"""Persistencia de las facturas confirmadas en un .xlsx mensual con el layout oficial del 606."""
from pathlib import Path

from openpyxl import Workbook, load_workbook

import config


def ruta_archivo_mes(empresa_rnc: str, periodo: str) -> Path:
    carpeta = config.CARPETA_DATOS / empresa_rnc
    carpeta.mkdir(parents=True, exist_ok=True)
    return carpeta / f"606_{periodo}.xlsx"


def _abrir_o_crear(empresa_rnc: str, periodo: str):
    ruta = ruta_archivo_mes(empresa_rnc, periodo)
    if ruta.exists():
        wb = load_workbook(ruta)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Detalle"
        ws.append(config.COLUMNAS_606)
    return wb, ws, ruta


def cargar_facturas(empresa_rnc: str, periodo: str) -> list[dict]:
    """Devuelve las facturas ya guardadas del período (para chequear NCF duplicados)."""
    ruta = ruta_archivo_mes(empresa_rnc, periodo)
    if not ruta.exists():
        return []
    wb = load_workbook(ruta)
    ws = wb.active
    encabezados = [c.value for c in ws[1]]
    facturas = []
    for fila in ws.iter_rows(min_row=2, values_only=True):
        if fila[1] is None:
            continue
        registro = dict(zip(encabezados, fila))
        facturas.append(
            {
                "ncf": registro.get("NCF"),
                "rnc_cedula": registro.get("RNC o Cedula"),
            }
        )
    return facturas


def agregar_factura(empresa_rnc: str, periodo: str, datos: dict) -> int:
    """Agrega una fila confirmada al Excel del período de esa empresa. Devuelve el número de línea asignado."""
    wb, ws, ruta = _abrir_o_crear(empresa_rnc, periodo)
    linea = ws.max_row  # fila 1 es encabezado; max_row ya apunta a la próxima línea de datos
    monto_bienes = datos["monto_bienes"]
    monto_servicios = datos["monto_servicios"]
    total = monto_bienes + monto_servicios
    tipo_bien_cod = datos["tipo_bien_servicio_sugerido"]
    forma_pago_cod = datos["forma_pago_sugerida"]
    fila = [
        linea,
        datos["rnc_cedula"],
        int(datos["tipo_id"]),
        f'{tipo_bien_cod}-{config.TIPO_BIENES_SERVICIOS.get(tipo_bien_cod, "")}',
        datos["ncf"],
        datos.get("ncf_modificado") or "",
        datos["fecha_comprobante"],
        datos["fecha_pago"],
        monto_servicios,
        monto_bienes,
        total,
        datos["itbis_facturado"],
        0,  # ITBIS Retenido
        0,  # ITBIS sujeto a Proporcionalidad
        0,  # ITBIS llevado al Costo
        # ITBIS por Adelantar: la herramienta oficial de DGII lo calcula como
        # ITBIS Facturado - ITBIS llevado al Costo (aquí el costo siempre es 0).
        datos["itbis_facturado"],
        0,  # ITBIS percibido en compras
        "",  # Tipo de Retencion en ISR
        0,  # Monto Retencion Renta
        0,  # ISR Percibido en compras
        0,  # Impuesto Selectivo al Consumo
        0,  # Otros Impuesto/Tasas
        datos.get("monto_propina_legal", 0),
        f'{forma_pago_cod}-{config.FORMA_PAGO.get(forma_pago_cod, "")}',
    ]
    ws.append(fila)
    wb.save(ruta)
    return linea
