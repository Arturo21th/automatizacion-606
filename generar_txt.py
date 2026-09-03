"""Genera el archivo TXT de envío del Formato 606 a partir del Excel mensual del período.

El layout replica el que genera la "Herramienta Formato 606" oficial de DGII
(versión 2025, macro GenerarArchivo): encabezado "606|RNC|período|cantidad",
23 campos de detalle separados por "|", fechas como AAAAMMDD y códigos de
catálogo de 2 dígitos. Aun así, conviene pasar el archivo por la Herramienta
de Prevalidación de DGII antes de un envío real.
"""
import sys
from pathlib import Path

from openpyxl import load_workbook

import almacen
import config


def generar_txt(empresa_rnc: str, periodo: str) -> Path:
    ruta_xlsx = almacen.ruta_archivo_mes(empresa_rnc, periodo)
    if not ruta_xlsx.exists():
        raise FileNotFoundError(f"No existe el Excel del período {periodo}: {ruta_xlsx}")

    wb = load_workbook(ruta_xlsx)
    ws = wb.active

    lineas = []
    for fila in ws.iter_rows(min_row=2, values_only=True):
        if fila[1] is None:
            continue
        (
            _linea, rnc, tipo_id, tipo_bien, ncf, ncf_mod, fecha_comp, fecha_pago,
            monto_serv, monto_bienes, total, itbis_fact, itbis_ret, itbis_prop,
            itbis_costo, itbis_adelantar, itbis_percibido, tipo_ret_isr,
            monto_ret_renta, isr_percibido, isc, otros_impuestos, propina, forma_pago,
        ) = fila

        campos = [
            rnc,
            tipo_id,
            str(tipo_bien).split("-")[0].strip(),
            ncf,
            ncf_mod or "",
            fecha_comp,
            fecha_pago,
            monto_serv or 0,
            monto_bienes or 0,
            total or 0,
            itbis_fact or 0,
            itbis_ret or 0,
            itbis_prop or 0,
            itbis_costo or 0,
            itbis_adelantar or 0,
            itbis_percibido or 0,
            str(tipo_ret_isr).split("-")[0].strip() if tipo_ret_isr else "",
            monto_ret_renta or 0,
            isr_percibido or 0,
            isc or 0,
            otros_impuestos or 0,
            propina or 0,
            str(forma_pago).split("-")[0].strip(),
        ]
        lineas.append("|".join(str(c) for c in campos))

    # Línea de encabezado del formato de envío: 606|RNC del declarante|período|cantidad de registros
    encabezado = f"606|{empresa_rnc}|{periodo}|{len(lineas)}"

    # Mismo nombre de archivo que produce la herramienta oficial de DGII.
    ruta_txt = config.CARPETA_DATOS / empresa_rnc / f"DGII_F_606_{empresa_rnc}_{periodo}.TXT"
    ruta_txt.write_text("\n".join([encabezado] + lineas), encoding="utf-8")
    return ruta_txt


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Uso: python generar_txt.py RNC_EMPRESA AAAAMM")
        sys.exit(1)
    ruta = generar_txt(sys.argv[1], sys.argv[2])
    print(f"Generado: {ruta}")
